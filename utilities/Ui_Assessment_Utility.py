import configparser
import json
from time import process_time_ns

import openpyxl


class Utility_Class:

    @staticmethod
    def read_Config_Details():
        config_details = configparser.ConfigParser()
        config_details.read('../testdata/Properties.ini')
        # ../testdata/Properties.ini
        # D:/Learn/2024/PythonRelated/Python_UI_Assessment/testdata/Properties.ini
        return  config_details

    @staticmethod
    def read_Json_File():

        with open('../testdata/testDataJson.json') as jsonFile:
            jsonFileDict = json.load(jsonFile)
            return jsonFileDict

    @staticmethod
    def read_Excel_File(filePath):
        work_Book = openpyxl.load_workbook(filePath)
        work_Sheet = work_Book.active
        test_Data_Dict={}
        for i in range(1,work_Sheet.max_row+1):
            for j in range(1,work_Sheet.max_column+1):
                test_Data_Dict[work_Sheet.cell(row=1,column=j).value] = work_Sheet.cell(row=i,column=j).value
                #print(work_Sheet.cell(row=1,column=j).value)
                #print(work_Sheet.cell(row=i,column=j).value, sep= ' ')
        return test_Data_Dict
